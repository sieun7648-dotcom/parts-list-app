import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import io

st.set_page_config(
    page_title="ROBOSTAR PARTS LIST 생성기",
    page_icon="🤖",
    layout="wide",
)

st.markdown("""
<style>
  [data-testid="stAppViewContainer"] { background: #F7F7F7; }
  [data-testid="stHeader"] { background: #fff; }
  .main-title {
    background: #fff; color: #111;
    padding: 14px 0 10px 0;
    font-size: 16px; font-weight: 600;
    margin-bottom: 20px; border-bottom: 2px solid #111;
    display: flex; align-items: center; gap: 10px;
  }
  .main-title span { color: #888; font-size: 12px; font-weight: 400; }
  .card {
    background: #fff; border: 1px solid #E0E0E0;
    border-radius: 8px; padding: 18px 20px; margin-bottom: 14px;
  }
  .card-title {
    font-size: 11px; font-weight: 600; color: #444;
    text-transform: uppercase; letter-spacing: .5px;
    margin-bottom: 12px; display: flex; align-items: center; gap: 8px;
  }
  .step-badge {
    background: #333; color: #fff;
    width: 20px; height: 20px; border-radius: 50%;
    display: inline-flex; align-items: center; justify-content: center;
    font-size: 10px; font-weight: 800;
  }
  .stButton > button {
    background: #222; color: #fff;
    border: none; border-radius: 5px;
    font-weight: 600; padding: 8px 20px;
    transition: background .15s;
  }
  .stButton > button:hover { background: #444; }
  .stTabs [data-baseweb="tab"] { font-size: 12px; }
  .stCheckbox label { font-size: 12px; }
</style>
""", unsafe_allow_html=True)

# ── 카테고리 ─────────────────────────────────────────────
CATS = [
    {"key": "BALL SCREW 가공코드", "pfx": "01002",   "type": "bscode"},
    {"key": "BALL SCREW",          "pfx": "03001",   "type": "normal"},
    {"key": "LM GUIDE",            "pfx": "02004",   "type": "normal"},
    {"key": "BALL BEARING",        "pfx": "07002",   "type": "normal"},
    {"key": "ANGULAR BEARING",     "pfx": "07001",   "type": "normal"},
    {"key": "PULLEY",              "pfx": "0900115", "type": "normal"},
    {"key": "TIMING BELT",         "pfx": "1300",    "type": "normal"},
    {"key": "COUPLING",            "pfx": "10002",   "type": "normal"},
    {"key": "POWER LOCK",          "pfx": "10004",   "type": "normal"},
    {"key": "SCALE HEAD",           "pfx": "06002",   "type": "normal"},
    {"key": "SCALE TAPE",           "pfx": "06005",   "type": "normal"},
    {"key": "LINEAR MOTOR",         "pfx": "05004",   "type": "normal"},
]

# ── 엑셀 스타일 상수 ─────────────────────────────────────
NAVY   = "333333"
NAVY2  = "555555"
WHITE  = "FFFFFF"
BLACK  = "111111"
LGRAY  = "F5F5F5"
STRIPE = "FAFAFA"
LINE   = "DDDDDD"

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(bold=False, color="111111", sz=10, name="맑은 고딕"):
    return Font(name=name, bold=bold, color=color, size=sz)
def align(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v)
def border():
    s = Side(style="thin", color=LINE)
    return Border(left=s, right=s, top=s, bottom=s)
def border_right():
    thin = Side(style="thin", color=LINE)
    med  = Side(style="medium", color="999999")
    return Border(left=thin, right=med, top=thin, bottom=thin)
def sc(cell, f=None, fl=None, al=None, bd=None):
    if f:  cell.font      = f
    if fl: cell.fill      = fl
    if al: cell.alignment = al
    if bd: cell.border    = bd

# ── BOM 읽기 (신규 ERP 파일) ─────────────────────────────
def read_bom_file(uploaded):
    wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
    ws = wb.active
    rows = [[str(c).strip() if c is not None else "" for c in row]
            for row in ws.iter_rows(values_only=True)]
    wb.close()
    col_map, header_row = {}, -1
    for i, row in enumerate(rows[:15]):
        if "제품품번" in row and "자재명" in row:
            header_row = i
            for j, v in enumerate(row):
                if v == "제품품번":  col_map["품번"]    = j
                if v == "자재명":   col_map["자재명"]  = j
                if v == "자재번호": col_map["자재번호"] = j
                if v in ("총소요량","자재소요량","소요량") and "수량" not in col_map:
                    col_map["수량"] = j
            break
    if header_row < 0:
        raise ValueError("제품품번·자재명 컬럼을 찾을 수 없습니다")
    data = []
    for row in rows[header_row+1:]:
        pn = row[col_map.get("품번", 0)]
        jm = row[col_map.get("자재명", 0)]
        jb = row[col_map.get("자재번호", 0)] if col_map.get("자재번호") is not None else ""
        try:    qty = float(row[col_map["수량"]]) if col_map.get("수량") is not None else 1
        except: qty = 1
        if pn and jm and jm != "TOTAL":
            data.append({"품번": pn, "자재번호": jb, "자재명": jm, "수량": qty})
    return data

# ── BOM 읽기 (구 R2P 붙여넣기) ──────────────────────────
def read_bom_paste(text):
    """
    구 R2P 복사 데이터 파싱
    구조: (빈탭) | 자재코드 | 자재명 | 도면번호 | 단위 | 품목구분 | 소요량 | ... | 제품코드
    """
    import re
    lines = [l for l in text.strip().splitlines() if l.strip()]
    if not lines:
        raise ValueError("데이터가 없습니다")

    def parse_line(line):
        cols = line.split("\t")
        if len(cols) < 4:
            return None
        pn = cols[-1].strip()
        if not pn:
            return None
        jb, jm, qty = "", "", 1
        code_idx = -1
        for i, c in enumerate(cols[:-1]):
            cs = c.strip()
            if cs and re.match(r"^[0-9A-Za-z]", cs):
                jb = cs
                code_idx = i
                break
        if code_idx < 0:
            return None
        for j in range(code_idx + 1, len(cols) - 1):
            cs = cols[j].strip()
            if not cs or cs == " ":
                continue
            if re.match(r"^(EA|MT|PC|SET|KG|R|M|L)$", cs, re.I):
                for k in range(j, min(j + 4, len(cols) - 1)):
                    try:
                        qty = float(cols[k].strip().replace(",", ""))
                        break
                    except:
                        pass
                break
            # 자재명: 한글 있거나, 콤마 있거나, 공백 있거나, 길이 > 12
            if re.search(r"[가-힣]", cs) or "," in cs or " " in cs or len(cs) > 12:
                jm = cs
                for k in range(j + 1, len(cols) - 1):
                    cs2 = cols[k].strip()
                    if re.match(r"^(EA|MT|PC|SET|KG|R|M|L)$", cs2, re.I):
                        for m in range(k, min(k + 4, len(cols) - 1)):
                            try:
                                qty = float(cols[m].strip().replace(",", ""))
                                break
                            except:
                                pass
                        break
                break
        if not jm:
            return None
        return {"품번": pn, "자재번호": jb, "자재명": jm, "수량": qty}

    header_kw = ["품번", "자재", "코드", "품명", "수량", "소요량", "단위"]
    data = []
    for line in lines:
        cols_check = line.split("\t")
        first_vals = " ".join(c.strip() for c in cols_check[:5])
        if any(k in first_vals for k in header_kw):
            continue
        row = parse_line(line)
        if row and row["품번"] and row["자재명"]:
            data.append(row)

    if not data:
        raise ValueError("파싱된 데이터가 없습니다. R2P에서 복사한 데이터인지 확인해주세요")
    return data


# ── 파츠리스트 빌드 ──────────────────────────────────────
def build_parts_list(bom, selected_cats):
    products = list(dict.fromkeys(r["품번"] for r in bom))
    result = []
    for prod in products:
        sub = [r for r in bom if r["품번"] == prod]
        row = {"품번": prod}
        for cat in selected_cats:
            matched = [r for r in sub if r["자재번호"].startswith(cat["pfx"])]
            if cat["type"] == "bscode":
                row[cat["key"]] = [{"code": r["자재번호"]} for r in matched]
            else:
                row[cat["key"]] = [{"model": r["자재명"], "qty": r["수량"]} for r in matched]
        if any(row[c["key"]] for c in selected_cats):
            result.append(row)
    return result

# ── 엑셀 생성 ────────────────────────────────────────────
def make_excel(result, selected_cats):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 가로형
    ws1 = wb.create_sheet("가로형")
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 38
    col_idx = 2
    cat_col = {}
    for cat in selected_cats:
        if cat["type"] == "bscode":
            ws1.column_dimensions[get_column_letter(col_idx)].width = 18
            col_idx += 1
        else:
            ws1.column_dimensions[get_column_letter(col_idx)].width = 38
            ws1.column_dimensions[get_column_letter(col_idx+1)].width = 7
            col_idx += 2

    ws1.row_dimensions[1].height = 22
    ws1.row_dimensions[2].height = 16

    c = ws1.cell(1, 1, "제품품번")
    sc(c, font(True,WHITE,10), fill(NAVY), align("center"), border())

    col_idx = 2
    for cat in selected_cats:
        cols = 1 if cat["type"] == "bscode" else 2
        c = ws1.cell(1, col_idx, cat["key"])
        sc(c, font(True,WHITE,10), fill(NAVY), align("center"), border())
        if cols == 2:
            ws1.merge_cells(start_row=1,start_column=col_idx,end_row=1,end_column=col_idx+1)
        if cat["type"] == "bscode":
            c2 = ws1.cell(2, col_idx, "가공코드")
        else:
            c2 = ws1.cell(2, col_idx, "자재명")
            c3 = ws1.cell(2, col_idx+1, "수량")
            sc(c3, font(True,"EEEEEE",9), fill(NAVY2), align("center"), border())
        sc(c2, font(True,"EEEEEE",9), fill(NAVY2), align("center"), border())
        cat_col[cat["key"]] = col_idx
        col_idx += cols

    ws1.merge_cells("A1:A2")
    ws1.cell(2,1).fill   = fill(NAVY)
    ws1.cell(2,1).border = border()

    row_idx = 3
    for pi, row in enumerate(result):
        is_alt = pi % 2 == 1
        row_fill = fill(STRIPE) if is_alt else fill(WHITE)
        max_n = max(1, max(len(row[c["key"]]) for c in selected_cats))
        ws1.row_dimensions[row_idx].height = 15
        c = ws1.cell(row_idx, 1, row["품번"])
        sc(c, font(True,BLACK), fill(LGRAY), align("center"), border_right())
        if max_n > 1:
            ws1.merge_cells(start_row=row_idx,start_column=1,end_row=row_idx+max_n-1,end_column=1)
        for cat in selected_cats:
            items = row[cat["key"]]
            s = cat_col[cat["key"]]
            cols = 1 if cat["type"] == "bscode" else 2
            if not items:
                c = ws1.cell(row_idx, s, "")
                sc(c, font(color="999999"), fill(WHITE), align("center"), border())
                if max_n > 1 or cols > 1:
                    ws1.merge_cells(start_row=row_idx,start_column=s,end_row=row_idx+max_n-1,end_column=s+cols-1)
            else:
                for i, item in enumerate(items):
                    r = row_idx+i
                    if i > 0: ws1.row_dimensions[r].height = 15
                    if cat["type"] == "bscode":
                        c = ws1.cell(r, s, item["code"])
                        sc(c, font(bold=True,color="333333"), row_fill, align("left"), border())
                    else:
                        cm = ws1.cell(r, s, item["model"])
                        cq = ws1.cell(r, s+1, item["qty"])
                        sc(cm, font(color=BLACK), row_fill, align("left"), border())
                        sc(cq, font(bold=True,color=BLACK), row_fill, align("center"), border())
                for i in range(len(items), max_n):
                    r = row_idx+i
                    ws1.row_dimensions[r].height = 15
                    for x in range(cols):
                        ws1.cell(r, s+x).fill   = row_fill
                        ws1.cell(r, s+x).border = border()
        row_idx += max_n
    ws1.freeze_panes = "A3"

    # 세로형
    ws2 = wb.create_sheet("세로형")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 38
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 42
    ws2.column_dimensions["E"].width = 7
    ws2.row_dimensions[1].height = 20

    for ci, v in enumerate(["No","제품품번","카테고리","자재명","수량"], 1):
        c = ws2.cell(1, ci, v)
        sc(c, font(True,WHITE,10), fill(NAVY), align("center"), border())

    r2, no = 2, 1
    for pi, row in enumerate(result):
        is_alt = pi % 2 == 1
        row_fill = fill(STRIPE) if is_alt else fill(WHITE)
        lines = []
        for cat in selected_cats:
            for item in row[cat["key"]]:
                if cat["type"] == "bscode":
                    lines.append((cat["key"], item["code"], ""))
                else:
                    lines.append((cat["key"], item["model"], item["qty"]))
        if not lines: continue
        cn = ws2.cell(r2, 1, no)
        sc(cn, font(bold=True,color=BLACK), fill(LGRAY), align("center"), border())
        if len(lines) > 1:
            ws2.merge_cells(start_row=r2,start_column=1,end_row=r2+len(lines)-1,end_column=1)
        cp = ws2.cell(r2, 2, row["품번"])
        sc(cp, font(bold=True,color=BLACK), fill(LGRAY), align("center"), border_right())
        if len(lines) > 1:
            ws2.merge_cells(start_row=r2,start_column=2,end_row=r2+len(lines)-1,end_column=2)
        for i, (cat_name, model, qty) in enumerate(lines):
            ws2.row_dimensions[r2+i].height = 15
            sc(ws2.cell(r2+i,3,cat_name), font(color=BLACK), row_fill, align("left"),   border())
            sc(ws2.cell(r2+i,4,model),    font(color=BLACK), row_fill, align("left"),   border())
            sc(ws2.cell(r2+i,5,qty),      font(bold=True,color=BLACK), row_fill, align("center"), border())
        r2 += len(lines)
        no += 1
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── UI ───────────────────────────────────────────────────
st.markdown("""
<div class="main-title">
  <span style="font-size:18px">●</span> ROBOSTAR PARTS LIST 생성기
  <span>다중제품 BOM → 파츠리스트 자동 변환</span>
</div>
""", unsafe_allow_html=True)

col1, col2 = st.columns([1, 1], gap="large")

with col1:
    # STEP 1 - 파일 입력
    st.markdown('<div class="card"><div class="card-title"><span class="step-badge">1</span> BOM 입력</div>', unsafe_allow_html=True)
    tab1, tab2 = st.tabs(["📁 신규 ERP (파일 업로드)", "📋 구 R2P (붙여넣기)"])

    bom = None
    with tab1:
        uploaded = st.file_uploader("BOM 파일", type=["xlsx","xlsm","xls"], label_visibility="collapsed")
        if uploaded:
            try:
                bom = read_bom_file(uploaded)
                prods = list(dict.fromkeys(r["품번"] for r in bom))
                st.success(f"✅ {uploaded.name} — 제품 {len(prods)}종 · 자재 {len(bom)}건")
                counts = {c["key"]: sum(1 for r in bom if r["자재번호"].startswith(c["pfx"])) for c in CATS}
            except Exception as e:
                st.error(f"❌ {e}")

    with tab2:
        st.caption("구 R2P에서 제품코드 / 자재코드 / 자재명 / 수량 컬럼을 복사해서 붙여넣기 하세요")
        paste_text = st.text_area("여기에 붙여넣기", height=160, placeholder="제품코드\t자재코드\t자재명\t수량\nRS12A...\t01002...\tBALL SCREW...\t1", label_visibility="collapsed")
        if st.button("파싱하기", key="parse_btn"):
            if paste_text.strip():
                try:
                    bom = read_bom_paste(paste_text)
                    prods = list(dict.fromkeys(r["품번"] for r in bom))
                    st.success(f"✅ 제품 {len(prods)}종 · 자재 {len(bom)}건 파싱 완료")
                    counts = {c["key"]: sum(1 for r in bom if r["자재번호"].startswith(c["pfx"])) for c in CATS}
                    st.session_state["bom"] = bom
                    st.session_state["counts"] = counts
                except Exception as e:
                    st.error(f"❌ {e}")
            else:
                st.warning("내용을 붙여넣어 주세요")

    # session_state에서 복원 (붙여넣기 후 다른 조작 시 유지)
    if bom is None and "bom" in st.session_state:
        bom = st.session_state["bom"]
        counts = st.session_state.get("counts", {})

    st.markdown('</div>', unsafe_allow_html=True)

    # STEP 2 - 자재 선택
    st.markdown('<div class="card"><div class="card-title"><span class="step-badge">2</span> 포함할 자재 선택</div>', unsafe_allow_html=True)

    # 전체선택/전체해제
    cb1, cb2, _ = st.columns([1, 1, 4])
    with cb1:
        if st.button("전체선택", key="all_on"):
            for cat in CATS:
                st.session_state[f"chk_{cat['key']}"] = True
    with cb2:
        if st.button("전체해제", key="all_off"):
            for cat in CATS:
                st.session_state[f"chk_{cat['key']}"] = False

    selected = {}
    cols_chk = st.columns(2)
    for i, cat in enumerate(CATS):
        cnt = counts.get(cat["key"], 0) if bom and "counts" in dir() else 0
        with cols_chk[i % 2]:
            selected[cat["key"]] = st.checkbox(
                f"{cat['key']} ({cnt}건)",
                value=st.session_state.get(f"chk_{cat['key']}", True),
                key=f"chk_{cat['key']}"
            )
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="card"><div class="card-title"><span class="step-badge">3</span> 파츠리스트 생성 & 다운로드</div>', unsafe_allow_html=True)

    selected_cats = [c for c in CATS if selected.get(c["key"])]

    if bom and selected_cats:
        result = build_parts_list(bom, selected_cats)

        # 미리보기
        preview_data = []
        for row in result:
            for cat in selected_cats:
                for item in row[cat["key"]]:
                    if cat["type"] == "bscode":
                        preview_data.append({"제품품번": row["품번"], "카테고리": cat["key"], "자재명": item["code"], "수량": ""})
                    else:
                        preview_data.append({"제품품번": row["품번"], "카테고리": cat["key"], "자재명": item["model"], "수량": item["qty"]})

        if preview_data:
            df = pd.DataFrame(preview_data)
            st.dataframe(df, use_container_width=True, height=420, hide_index=True)

        st.markdown("---")

        fname = f"PARTS_LIST_{date.today().strftime('%Y%m%d')}.xlsx"
        buf = make_excel(result, selected_cats)
        st.download_button(
            label="⬇ 엑셀 다운로드 (가로형 + 세로형)",
            data=buf,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.info("BOM을 입력하고 자재를 선택하면 미리보기가 생성됩니다.")

    st.markdown('</div>', unsafe_allow_html=True)
